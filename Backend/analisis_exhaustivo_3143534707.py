#!/usr/bin/env python3
"""
ANÁLISIS EXHAUSTIVO DEL NÚMERO 3143534707
========================================

Análisis específico para verificar presencia del número 3143534707
en los 4 archivos Excel de CLARO identificados por Boris.

Celdas objetivo según Boris: 53591, 51438, 56124, 51203

Autor: Claude Code - Data Engineering Algorithm Expert
Para: Boris - KRONOS Project
Fecha: 2025-08-18
"""

import pandas as pd
import openpyxl
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set, Tuple, Any
import sqlite3
import logging

# Configurar logging detallado
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Number3143534707Analyzer:
    """
    Analizador especializado para el número 3143534707
    
    Realiza búsqueda exhaustiva en archivos Excel específicos
    y verifica correlación con celdas HUNTER identificadas por Boris.
    """
    
    def __init__(self):
        self.target_number = '3143534707'
        self.expected_cells = ['53591', '51438', '56124', '51203']
        
        # Archivos específicos de CLARO mencionados por Boris
        self.excel_files = [
            r"C:\Soluciones\BGC\claude\KNSOft\archivos\envioarchivosparaanalizar (1)\1-225211_LLAMADAS_ENTRANTES_POR_CELDA_545612_0.xlsx",
            r"C:\Soluciones\BGC\claude\KNSOft\archivos\envioarchivosparaanalizar (1)\1-225211_LLAMADAS_SALIENTES_POR_CELDA_545613_0.xlsx",
            r"C:\Soluciones\BGC\claude\KNSOft\archivos\envioarchivosparaanalizar (1)\2-225211_LLAMADAS_ENTRANTES_POR_CELDA_545614_0.xlsx",
            r"C:\Soluciones\BGC\claude\KNSOft\archivos\envioarchivosparaanalizar (1)\2-225211_LLAMADAS_SALIENTES_POR_CELDA_545615_0.xlsx"
        ]
        
        # Patrones de búsqueda exhaustivos para el número
        self.number_patterns = [
            self.target_number,                    # 3143534707
            f"57{self.target_number}",            # 573143534707
            f"+57{self.target_number}",           # +573143534707
            f"57 {self.target_number}",           # 57 3143534707
            f"+57 {self.target_number}",          # +57 3143534707
            f"0057{self.target_number}",          # 00573143534707
        ]
        
        self.results = {
            'analysis_timestamp': datetime.now().isoformat(),
            'target_number': self.target_number,
            'expected_cells': self.expected_cells,
            'files_analyzed': [],
            'findings': [],
            'summary': {
                'total_occurrences': 0,
                'cells_found': [],
                'files_with_number': 0,
                'verification_status': 'PENDING'
            }
        }

    def analyze_excel_structure(self, file_path: str) -> Dict[str, Any]:
        """
        Analiza la estructura de un archivo Excel
        
        Args:
            file_path: Ruta al archivo Excel
            
        Returns:
            Dict con información de estructura del archivo
        """
        logger.info(f"Analizando estructura: {Path(file_path).name}")
        
        try:
            # Leer usando openpyxl para mejor control
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            file_info = {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'sheets': [],
                'total_rows': 0,
                'structure_analysis': True
            }
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Leer primera fila para headers
                headers = []
                if max_row > 0:
                    first_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    headers = [str(cell) if cell is not None else f"Col_{i}" for i, cell in enumerate(first_row)]
                
                sheet_info = {
                    'name': sheet_name,
                    'max_row': max_row,
                    'max_col': max_col,
                    'headers': headers,
                    'has_data': max_row > 1
                }
                
                file_info['sheets'].append(sheet_info)
                file_info['total_rows'] += max_row
                
                logger.info(f"  Hoja '{sheet_name}': {max_row} filas, {max_col} columnas")
                logger.info(f"  Headers: {headers[:5]}...")  # Primeros 5 headers
                
            workbook.close()
            return file_info
            
        except Exception as e:
            logger.error(f"Error analizando estructura de {file_path}: {str(e)}")
            return {
                'file_path': file_path,
                'file_name': Path(file_path).name,
                'error': str(e),
                'structure_analysis': False
            }

    def search_number_in_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Busca el número objetivo en todas las hojas de un archivo Excel
        
        Args:
            file_path: Ruta al archivo Excel
            
        Returns:
            Dict con resultados de la búsqueda
        """
        logger.info(f"Buscando número {self.target_number} en: {Path(file_path).name}")
        
        search_result = {
            'file_path': file_path,
            'file_name': Path(file_path).name,
            'number_found': False,
            'occurrences': [],
            'cells_with_number': set(),
            'total_matches': 0,
            'search_successful': True
        }
        
        try:
            # Usar pandas para búsqueda eficiente
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                logger.info(f"  Analizando hoja: {sheet_name}")
                
                try:
                    # Leer hoja completa
                    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                    
                    logger.info(f"    Dimensiones: {df.shape}")
                    logger.info(f"    Columnas: {list(df.columns)}")
                    
                    # Buscar en todas las columnas
                    for col_name in df.columns:
                        if pd.api.types.is_object_dtype(df[col_name]):
                            # Convertir a string y buscar patrones
                            col_str = df[col_name].astype(str)
                            
                            for pattern in self.number_patterns:
                                mask = col_str.str.contains(pattern, na=False, regex=False)
                                matches = df[mask]
                                
                                if not matches.empty:
                                    logger.info(f"    ¡ENCONTRADO! Patrón '{pattern}' en columna '{col_name}': {len(matches)} coincidencias")
                                    
                                    search_result['number_found'] = True
                                    
                                    for idx, row in matches.iterrows():
                                        # Buscar columna de celda
                                        cell_id = None
                                        for potential_cell_col in ['CELDA', 'CELL_ID', 'ID_CELDA', 'CODIGO_CELDA', 'CELDA_ID']:
                                            if potential_cell_col in df.columns:
                                                cell_id = str(row[potential_cell_col])
                                                break
                                        
                                        occurrence = {
                                            'sheet': sheet_name,
                                            'row': idx + 2,  # +2 porque pandas es 0-based y Excel 1-based + header
                                            'column': col_name,
                                            'pattern_found': pattern,
                                            'cell_value': str(row[col_name]),
                                            'cell_id': cell_id,
                                            'full_row_data': dict(row)
                                        }
                                        
                                        search_result['occurrences'].append(occurrence)
                                        
                                        if cell_id:
                                            search_result['cells_with_number'].add(cell_id)
                                            logger.info(f"      Celda encontrada: {cell_id}")
                    
                except Exception as e:
                    logger.error(f"    Error procesando hoja {sheet_name}: {str(e)}")
                    continue
            
            search_result['cells_with_number'] = list(search_result['cells_with_number'])
            search_result['total_matches'] = len(search_result['occurrences'])
            
            logger.info(f"  Resultado: {search_result['total_matches']} coincidencias, celdas: {search_result['cells_with_number']}")
            
        except Exception as e:
            logger.error(f"Error buscando en {file_path}: {str(e)}")
            search_result['search_successful'] = False
            search_result['error'] = str(e)
        
        return search_result

    def verify_expected_cells(self, found_cells: Set[str]) -> Dict[str, Any]:
        """
        Verifica si las celdas encontradas coinciden con las esperadas por Boris
        
        Args:
            found_cells: Set de celdas encontradas
            
        Returns:
            Dict con resultado de verificación
        """
        found_set = set(found_cells)
        expected_set = set(self.expected_cells)
        
        verification = {
            'expected_cells': self.expected_cells,
            'found_cells': list(found_set),
            'matching_cells': list(found_set.intersection(expected_set)),
            'missing_cells': list(expected_set - found_set),
            'extra_cells': list(found_set - expected_set),
            'exact_match': found_set == expected_set,
            'partial_match': bool(found_set.intersection(expected_set)),
            'match_percentage': len(found_set.intersection(expected_set)) / len(expected_set) * 100 if expected_set else 0
        }
        
        logger.info(f"Verificación de celdas:")
        logger.info(f"  Esperadas por Boris: {verification['expected_cells']}")
        logger.info(f"  Encontradas: {verification['found_cells']}")
        logger.info(f"  Coincidencias: {verification['matching_cells']}")
        logger.info(f"  Faltantes: {verification['missing_cells']}")
        logger.info(f"  Extra: {verification['extra_cells']}")
        logger.info(f"  Coincidencia: {verification['match_percentage']:.1f}%")
        
        return verification

    def check_database_correlation(self) -> Dict[str, Any]:
        """
        Verifica cómo está el número en la base de datos SQLite
        
        Returns:
            Dict con estado en BD
        """
        logger.info("Verificando estado en base de datos...")
        
        db_status = {
            'database_accessible': False,
            'operator_records': [],
            'cellular_records': [],
            'total_operator_matches': 0,
            'total_cellular_matches': 0,
            'cells_in_operator_data': set(),
            'verification_successful': True
        }
        
        try:
            # Conectar a base de datos KRONOS
            db_path = r"C:\Soluciones\BGC\claude\KNSOft\Backend\kronos.db"
            if not Path(db_path).exists():
                logger.warning(f"Base de datos no encontrada: {db_path}")
                db_status['verification_successful'] = False
                db_status['error'] = "Base de datos no accesible"
                return db_status
            
            with sqlite3.connect(db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                db_status['database_accessible'] = True
                
                # Buscar en operator_call_data (revisar todas las columnas de números)
                logger.info("  Buscando en tabla operator_call_data...")
                for pattern in self.number_patterns:
                    cursor.execute("""
                        SELECT mission_id, numero_origen, numero_destino, celda_origen, celda_destino, 
                               operator, fecha_hora_llamada, duracion_segundos
                        FROM operator_call_data 
                        WHERE numero_origen LIKE ? OR numero_origen = ? 
                           OR numero_destino LIKE ? OR numero_destino = ?
                    """, (f'%{pattern}%', pattern, f'%{pattern}%', pattern))
                    
                    records = cursor.fetchall()
                    for record in records:
                        record_dict = dict(record)
                        db_status['operator_records'].append(record_dict)
                        
                        # Agregar celdas encontradas
                        if record['celda_origen']:
                            db_status['cells_in_operator_data'].add(str(record['celda_origen']))
                        if record['celda_destino']:
                            db_status['cells_in_operator_data'].add(str(record['celda_destino']))
                
                db_status['total_operator_matches'] = len(db_status['operator_records'])
                
                # Buscar en cellular_data (HUNTER) para las celdas objetivo
                logger.info("  Buscando en tabla cellular_data...")
                cursor.execute("""
                    SELECT mission_id, cell_id, operator, tecnologia, lat, lon, rssi, created_at
                    FROM cellular_data 
                    WHERE cell_id IN ('53591', '51438', '56124', '51203')
                """)
                
                cellular_records = cursor.fetchall()
                db_status['cellular_records'] = [dict(record) for record in cellular_records]
                db_status['total_cellular_matches'] = len(db_status['cellular_records'])
                
                db_status['cells_in_operator_data'] = list(db_status['cells_in_operator_data'])
                
                logger.info(f"  Registros en operator_call_data: {db_status['total_operator_matches']}")
                logger.info(f"  Celdas en operator_data: {db_status['cells_in_operator_data']}")
                logger.info(f"  Registros HUNTER para celdas objetivo: {db_status['total_cellular_matches']}")
        
        except Exception as e:
            logger.error(f"Error verificando base de datos: {str(e)}")
            db_status['verification_successful'] = False
            db_status['error'] = str(e)
        
        return db_status

    def run_comprehensive_analysis(self) -> Dict[str, Any]:
        """
        Ejecuta análisis completo del número 3143534707
        
        Returns:
            Dict con todos los resultados del análisis
        """
        logger.info("=== INICIANDO ANÁLISIS EXHAUSTIVO DEL NÚMERO 3143534707 ===")
        
        # Verificar existencia de archivos
        available_files = []
        for file_path in self.excel_files:
            if Path(file_path).exists():
                available_files.append(file_path)
                logger.info(f"✓ Archivo disponible: {Path(file_path).name}")
            else:
                logger.warning(f"✗ Archivo no encontrado: {file_path}")
        
        if not available_files:
            logger.error("¡CRÍTICO! Ningún archivo Excel está disponible")
            return {
                'success': False,
                'error': 'No se encontraron archivos Excel para analizar',
                'timestamp': datetime.now().isoformat()
            }
        
        # 1. Analizar estructura de archivos
        logger.info("\n=== FASE 1: ANÁLISIS DE ESTRUCTURA ===")
        for file_path in available_files:
            structure_info = self.analyze_excel_structure(file_path)
            self.results['files_analyzed'].append(structure_info)
        
        # 2. Buscar número en cada archivo
        logger.info("\n=== FASE 2: BÚSQUEDA DEL NÚMERO ===")
        all_found_cells = set()
        files_with_number = 0
        
        for file_path in available_files:
            search_result = self.search_number_in_excel(file_path)
            self.results['findings'].append(search_result)
            
            if search_result['number_found']:
                files_with_number += 1
                all_found_cells.update(search_result['cells_with_number'])
                self.results['summary']['total_occurrences'] += search_result['total_matches']
        
        # 3. Verificar celdas esperadas
        logger.info("\n=== FASE 3: VERIFICACIÓN DE CELDAS ===")
        cell_verification = self.verify_expected_cells(all_found_cells)
        self.results['cell_verification'] = cell_verification
        
        # 4. Verificar estado en base de datos
        logger.info("\n=== FASE 4: VERIFICACIÓN DE BASE DE DATOS ===")
        db_status = self.check_database_correlation()
        self.results['database_status'] = db_status
        
        # 5. Resumen final
        self.results['summary'].update({
            'files_with_number': files_with_number,
            'cells_found': list(all_found_cells),
            'verification_status': 'VERIFIED' if cell_verification['exact_match'] else 'DISCREPANCY_DETECTED'
        })
        
        # 6. Generar diagnóstico
        diagnosis = self._generate_diagnosis()
        self.results['diagnosis'] = diagnosis
        
        logger.info("\n=== ANÁLISIS COMPLETADO ===")
        logger.info(f"Archivos con número: {files_with_number}/{len(available_files)}")
        logger.info(f"Total ocurrencias: {self.results['summary']['total_occurrences']}")
        logger.info(f"Celdas encontradas: {self.results['summary']['cells_found']}")
        logger.info(f"Estado: {self.results['summary']['verification_status']}")
        
        return self.results

    def _generate_diagnosis(self) -> Dict[str, Any]:
        """
        Genera diagnóstico detallado del análisis
        
        Returns:
            Dict con diagnóstico y recomendaciones
        """
        diagnosis = {
            'timestamp': datetime.now().isoformat(),
            'critical_findings': [],
            'recommendations': [],
            'algorithm_issues': [],
            'data_integrity_status': 'UNKNOWN'
        }
        
        # Analizar hallazgos críticos
        if self.results['summary']['total_occurrences'] == 0:
            diagnosis['critical_findings'].append("¡CRÍTICO! Número 3143534707 NO encontrado en ningún archivo Excel")
            diagnosis['recommendations'].append("Verificar que el número esté realmente en los archivos fuente")
            diagnosis['data_integrity_status'] = 'CORRUPTED_OR_MISSING'
        
        elif self.results['summary']['total_occurrences'] < 4:
            diagnosis['critical_findings'].append(f"DISCREPANCIA: Solo {self.results['summary']['total_occurrences']} ocurrencias encontradas, Boris reporta 4")
            diagnosis['algorithm_issues'].append("Posible falla en extracción de datos o criterios de búsqueda")
        
        # Verificar celdas
        cell_verification = self.results.get('cell_verification', {})
        if not cell_verification.get('exact_match', False):
            diagnosis['critical_findings'].append("Las celdas encontradas NO coinciden exactamente con las esperadas por Boris")
            diagnosis['algorithm_issues'].append("Falla en identificación correcta de celdas objetivo")
        
        # Verificar base de datos
        db_status = self.results.get('database_status', {})
        if db_status.get('total_operator_matches', 0) == 0:
            diagnosis['critical_findings'].append("Número NO encontrado en base de datos operator_call_data")
            diagnosis['recommendations'].append("Verificar proceso de carga de datos CLARO a SQLite")
        
        # Determinar causa raíz probable
        if self.results['summary']['total_occurrences'] > 0 and db_status.get('total_operator_matches', 0) == 0:
            diagnosis['algorithm_issues'].append("CAUSA RAÍZ: Datos presentes en Excel pero ausentes en BD - Falla en proceso ETL")
        elif self.results['summary']['total_occurrences'] == 0:
            diagnosis['algorithm_issues'].append("CAUSA RAÍZ: Datos ausentes en archivos fuente - Verificar archivos originales")
        
        return diagnosis

    def save_results(self, output_file: str = None) -> str:
        """
        Guarda resultados del análisis en archivo JSON
        
        Args:
            output_file: Ruta del archivo de salida (opcional)
            
        Returns:
            Ruta del archivo guardado
        """
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"analisis_3143534707_exhaustivo_{timestamp}.json"
        
        # Convertir sets a listas para serialización JSON
        json_results = self._convert_sets_to_lists(self.results)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"Resultados guardados en: {output_file}")
        return output_file

    def _convert_sets_to_lists(self, obj):
        """Convierte sets a listas recursivamente para serialización JSON"""
        if isinstance(obj, set):
            return list(obj)
        elif isinstance(obj, dict):
            return {key: self._convert_sets_to_lists(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._convert_sets_to_lists(item) for item in obj]
        else:
            return obj


def main():
    """Función principal para ejecutar el análisis"""
    try:
        analyzer = Number3143534707Analyzer()
        results = analyzer.run_comprehensive_analysis()
        
        # Guardar resultados
        output_file = analyzer.save_results()
        
        # Mostrar resumen ejecutivo
        print("\n" + "="*60)
        print("RESUMEN EJECUTIVO - ANÁLISIS 3143534707")
        print("="*60)
        print(f"Timestamp: {results['analysis_timestamp']}")
        print(f"Archivos analizados: {len(results['files_analyzed'])}")
        print(f"Ocurrencias totales: {results['summary']['total_occurrences']}")
        print(f"Archivos con número: {results['summary']['files_with_number']}")
        print(f"Celdas encontradas: {results['summary']['cells_found']}")
        print(f"Estado verificación: {results['summary']['verification_status']}")
        
        if 'diagnosis' in results:
            print(f"\nHALLAZGOS CRÍTICOS:")
            for finding in results['diagnosis']['critical_findings']:
                print(f"  ⚠️  {finding}")
        
        print(f"\nResultados completos en: {output_file}")
        
        return results
        
    except Exception as e:
        logger.error(f"Error durante análisis: {str(e)}")
        raise


if __name__ == "__main__":
    main()